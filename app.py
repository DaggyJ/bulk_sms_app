# app.py (rewritten and updated)
from flask import Flask, render_template, request, session, jsonify
from config import SECRET_KEY, CELCOM_API_KEY
from models import db, User, Contact, SMSLog
from utils import send_bulk_sms, get_celcom_balance
from werkzeug.security import check_password_hash, generate_password_hash
from openpyxl import load_workbook
import os
from sqlalchemy import func  

# -------------------------
# APP INITIALIZATION
# -------------------------
app = Flask(__name__)
app.secret_key = SECRET_KEY
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = "sqlite:///" + os.path.join(BASE_DIR, "instance", "contacts.db")
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

#Enable Button
@app.route("/enable_user", methods=["POST"])
def enable_user():
    if not session.get("is_admin"):
        return jsonify({"status": "Access denied"}), 403

    data = request.get_json()
    user_id = data.get("user_id")
    user = User.query.get(user_id)

    if not user:
        return jsonify({"status": "User not found"}), 404

    user.status = "approved"
    db.session.commit()
    return jsonify({"status": "User enabled successfully"})



# -------------------------
# SPA MAIN PAGE
# -------------------------
@app.route("/")
def index():
    return render_template("index.html")

# -------------------------
# LOGIN
# -------------------------
@app.route("/login", methods=["POST"])
def login():
    data = request.get_json() or {}
    username = data.get("username")
    password = data.get("password")
    rest_pin = data.get("rest_pin")

    user = User.query.filter_by(username=username).first()
    if not user or not check_password_hash(user.password_hash, password):
        return jsonify({"error": "Invalid credentials"}), 401
    if user.status == "pending":
        return jsonify({"error": "Your account is awaiting admin approval"}), 403
    if user.status == "disabled":
        return jsonify({"error": "Your account has been disabled"}), 403
    if rest_pin and str(user.rest_pin) != str(rest_pin):
        return jsonify({"error": "Invalid REST PIN"}), 401
    
    # Persist session
    session.permanent = True
    session['user_id'] = user.id
    session['is_admin'] = bool(user.is_admin)

    # Fetch Celcom balance for admins
    balance = 0
    if user.is_admin:
        try:
            balance = int(get_celcom_balance() or 0)
        except Exception:
            balance = 0
    session['balance'] = balance

    return jsonify({
        "status": "logged_in",
        "is_admin": user.is_admin,
        "isAdmin": user.is_admin,
        "balance": balance,
        "is_logged_in": True,
        "isLoggedIn": True
    })

# -------------------------
# LOGOUT
# -------------------------
@app.route("/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"status": "logged_out"})

# -------------------------
# REGISTER NEW USER
# -------------------------
@app.route("/register", methods=["POST"])
def register():
    data = request.get_json() or {}
    username = data.get("username")
    password = data.get("password")
    rest_pin = data.get("rest_pin")

    if not username or not password:
        return jsonify({"status": "Username and password required"}), 400
    if User.query.filter_by(username=username).first():
        return jsonify({"status": "User already exists"}), 400

    new_user = User(
        username=username,
        password_hash=generate_password_hash(password),
        rest_pin=rest_pin,
        is_admin=False,
        status="pending"
    )
    db.session.add(new_user)
    db.session.commit()
    return jsonify({"status": "Registration sent for approval"})

# -------------------------
# ADMIN USER MANAGEMENT
# -------------------------
@app.route("/get_all_users")
def get_all_users():
    if not session.get("is_admin"):
        return jsonify({"users": [], "status": "Access denied"}), 403
    users = User.query.all()
    users_data = [{"id": u.id, "username": u.username, "status": u.status} for u in users]
    return jsonify({"users": users_data})

@app.route("/approve_user", methods=["POST"])
def approve_user():
    if not session.get("is_admin"):
        return jsonify({"status": "Access denied"}), 403
    user_id = request.json.get("user_id")
    user = User.query.get(user_id)
    if not user:
        return jsonify({"status": "User not found"}), 404
    user.status = "approved"
    db.session.commit()
    return jsonify({"status": "User approved"})

@app.route("/reject_user", methods=["POST"])
def reject_user():
    if not session.get("is_admin"):
        return jsonify({"status": "Access denied"}), 403
    user_id = request.json.get("user_id")
    user = User.query.get(user_id)
    if not user:
        return jsonify({"status": "User not found"}), 404
    db.session.delete(user)
    db.session.commit()
    return jsonify({"status": "User rejected and removed"})

@app.route("/disable_user", methods=["POST"])
def disable_user():
    # Only admin can perform this action
    if not session.get("is_admin"):
        return jsonify({"status": "Access denied"}), 403

    data = request.get_json()
    user_id = data.get("user_id")

    user = User.query.get(user_id)
    if not user:
        return jsonify({"status": "User not found"}), 404

    # 🚫 Prevent disabling an admin account
    if user.is_admin:
        return jsonify({"status": "Cannot disable admin"}), 400

    # Disable the account
    user.status = "disabled"
    db.session.commit()

    return jsonify({"status": "User disabled successfully"})


# -------------------------
# USER STATUS & CELCOM BALANCE
# -------------------------
@app.route("/user_status")
def user_status():
    is_logged_in = 'user_id' in session
    is_admin = bool(session.get('is_admin', False))
    balance = 0

    # Try to get balance from session first
    try:
        balance = int(session.get('balance', 0) or 0)
    except (ValueError, TypeError):
        balance = 0

    # Fetch live balance if admin
    if is_logged_in and is_admin:
        try:
            celcom_balance = get_celcom_balance()
            # Convert to integer safely
            balance = int(float(celcom_balance) if celcom_balance else 0)
            session['balance'] = balance
        except Exception as e:
            app.logger.warning("get_celcom_balance failed: %s", e)
            # Keep previous balance or 0
            balance = session.get('balance', 0)

    return jsonify({
        "is_logged_in": is_logged_in,
        "isLoggedIn": is_logged_in,
        "is_admin": is_admin,
        "isAdmin": is_admin,
        "balance": balance
    })


# -------------------------
# CONTACTS
# -------------------------
@app.route("/get_contacts")
def get_contacts():
    category = request.args.get("category", "").strip()
    query = Contact.query
    if category:
        query = query.filter(func.lower(func.trim(Contact.category)) == category.lower())
    contacts = query.all()
    data = [
        {
            "id": c.id,
            "name": c.name,
            "phone": c.phone,
            "category": c.category,
            "region": c.region,
            "subregion": c.subregion
        }
        for c in contacts
    ]
    return jsonify({"contacts": data})

@app.route("/upload_contacts", methods=["POST"])
def upload_contacts():
    if not session.get("is_admin"):
        return jsonify({"status": "Access denied"}), 403
    file = request.files.get("file")
    if not file:
        return jsonify({"status": "No file uploaded"}), 400
    wb = load_workbook(file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, phone, category, region, subregion = row
        if not name or not phone or not category:
            continue
        contact = Contact(
            name=name.strip(),
            phone=str(phone).strip(),
            category=category.strip(),
            region=region.strip() if region else '',
            subregion=subregion.strip() if subregion else ''
        )
        db.session.add(contact)
    db.session.commit()
    return jsonify({"status": "Contacts uploaded successfully"})

# -------------------------
# SEND SMS & LOG
# -------------------------
@app.route("/send_sms", methods=["POST"])
def send_sms():
    data = request.get_json() or {}
    message = data.get("message")
    recipients = data.get("recipients")
    category = data.get("category", "")

    if not message or not recipients:
        return jsonify({"status": "Message and recipients are required"}), 400

    response = send_bulk_sms(message, recipients)

    # Log SMS
    log = SMSLog(
        message=message,
        recipients=str(recipients),
        category=category,
        api_response=str(response)
    )
    db.session.add(log)
    db.session.commit()

    return jsonify({"status": "OK", "provider_response": response})

# -------------------------
# SMS LOGS
# -------------------------
@app.route("/logs")
def get_logs():
    logs = SMSLog.query.order_by(SMSLog.sent_at.desc()).all()
    logs_data = [
        {"id": l.id, "message": l.message, "recipients": l.recipients, "sent_at": str(l.sent_at)}
        for l in logs
    ]
    return jsonify({"logs": logs_data})

# -------------------------
# RUN APP
# -------------------------
if __name__ == "__main__":
    with app.app_context():
        db.create_all()
    app.run(debug=True)


