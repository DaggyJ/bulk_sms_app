# seed_contacts.py
import os
from openpyxl import load_workbook
from app import app, db
from models import Contact

# Path to your Excel file
EXCEL_FILE = os.path.join(os.path.dirname(__file__), "contacts.xlsx")

def seed_contacts():
    if not os.path.exists(EXCEL_FILE):
        print(f"Excel file not found at {EXCEL_FILE}")
        return

    wb = load_workbook(EXCEL_FILE)
    sheet = wb.active

    with app.app_context():
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name, phone, category, region, subregion = row
            if not name or not phone or not category:
                continue

            contact = Contact(
                name=name.strip(),
                phone=str(phone).strip(),
                category=category.strip(),
                region=region.strip() if region else '',
                subregion=subregion.strip() if subregion else ''
            )
            db.session.add(contact)

        db.session.commit()
        print("Contacts have been successfully uploaded!")

if __name__ == "__main__":
    seed_contacts()
